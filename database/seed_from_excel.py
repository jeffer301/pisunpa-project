import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Dict, List, Tuple

import psycopg2
from psycopg2.extras import DictCursor

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
    raise SystemExit(
        "Falta la dependencia openpyxl. Instala las dependencias con: pip install -r database/requirements.txt"
    ) from exc


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL = ROOT / "database" / "Datos.xlsx"
DEFAULT_SQL_OUTPUT = ROOT / "database" / "seed_data.sql"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text.strip())
    return text


def canonical_key(value: object) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_name(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    parts = [part for part in text.split(" ") if part]
    return " ".join(part.capitalize() for part in parts)


def normalize_subject(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\bde\b", "de", text, flags=re.IGNORECASE)
    text = re.sub(r"\bdel\b", "del", text, flags=re.IGNORECASE)
    text = re.sub(r"\bla\b", "la", text, flags=re.IGNORECASE)
    return " ".join(word.capitalize() for word in text.split(" "))


def infer_jornada(group_code: str) -> str:
    return "nocturna" if re.search(r"[Nn]", group_code) else "diurna"


def infer_subject_code(subject_name: str, group_code: str) -> str:
    candidate = re.sub(r"[^A-Za-z0-9]+", "", subject_name).upper()
    if len(candidate) <= 4:
        candidate = re.sub(r"[^A-Za-z0-9]+", "", group_code).upper()
    match = re.match(r"^([A-Za-z]+\d+)", group_code.replace("-", ""))
    if match:
        return match.group(1).upper()
    if candidate:
        return candidate[:20]
    return "SIN_CODIGO"


def read_rows(excel_path: Path) -> List[Tuple[str, str, str, str]]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    rows: List[Tuple[str, str, str, str]] = []
    header_row = None
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if header_row is None and any(cell is not None and str(cell).strip() for cell in row):
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            lowered = [value.lower() for value in values]
            if any("docente" in value for value in lowered) or any("asignatura" in value for value in lowered):
                header_row = idx
                continue

    if header_row is None:
        raise ValueError("No se encontró una fila de encabezado en el Excel")

    for row in list(sheet.iter_rows(values_only=True))[header_row + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if len(values) < 6:
            values = values + [""] * (6 - len(values))

        teacher = values[1] if len(values) > 1 else ""
        email = values[2] if len(values) > 2 else ""
        subject = values[4] if len(values) > 4 else ""
        group_code = values[5] if len(values) > 5 else ""

        if not teacher and not email and not subject and not group_code:
            continue

        rows.append((teacher, email, subject, group_code))

    return rows


def build_seed_payload(rows: List[Tuple[str, str, str, str]]) -> Dict[str, List[Dict[str, object]]]:
    teachers: Dict[str, Dict[str, object]] = {}
    subjects: Dict[str, Dict[str, object]] = {}
    assignments: List[Dict[str, object]] = []

    for teacher, email, subject_name, group_code in rows:
        teacher_name = normalize_name(teacher)
        email_value = email.lower().strip() if email else ""
        subject_display = normalize_subject(subject_name)
        subject_key = canonical_key(subject_display)
        group_value = normalize_text(group_code).upper()

        if not email_value:
            continue

        if email_value not in teachers:
            teachers[email_value] = {
                "email": email_value,
                "nombres": teacher_name.split(" ", 1)[0] if teacher_name else "",
                "apellidos": teacher_name.split(" ", 1)[1] if " " in teacher_name else "",
                "nombre_completo": teacher_name,
            }

        if subject_key not in subjects:
            subjects[subject_key] = {
                "codigo": infer_subject_code(subject_display, group_value),
                "nombre": subject_display or "Sin nombre",
            }

        assignments.append(
            {
                "email": email_value,
                "subject_key": subject_key,
                "subject_name": subject_display,
                "group_code": group_value,
                "jornada": infer_jornada(group_value),
            }
        )

    return {
        "teachers": list(teachers.values()),
        "subjects": list(subjects.values()),
        "assignments": assignments,
    }


def write_sql_output(payload: Dict[str, List[Dict[str, object]]], output_path: Path) -> None:
    lines: List[str] = []
    lines.append("BEGIN;")
    lines.append("")
    lines.append("INSERT INTO roles (codigo, nombre, descripcion) VALUES")
    lines.append("    ('docente', 'Docente', 'Acceso a funciones academicas autorizadas.')")
    lines.append("ON CONFLICT (codigo) DO NOTHING;")
    lines.append("")
    lines.append("INSERT INTO programas_academicos (nombre, codigo, nivel, modalidad, sede)")
    lines.append("VALUES ('Ingenieria de Sistemas', 'ING-SIS', 'pregrado', 'presencial', 'Buenaventura')")
    lines.append("ON CONFLICT (codigo) DO NOTHING;")
    lines.append("")

    lines.append("-- Usuarios")
    for teacher in payload["teachers"]:
        lines.append(
            "INSERT INTO usuarios (email, password_hash, nombres, apellidos, estado, created_at, updated_at) VALUES "
            f"('{teacher['email']}', 'seed-placeholder', '{teacher['nombres']}', '{teacher['apellidos']}', 'activo', now(), now()) "
            "ON CONFLICT (email) DO NOTHING;"
        )
    lines.append("")

    lines.append("-- Asignar rol docente")
    lines.append(
        "INSERT INTO usuario_roles (usuario_id, rol_id, asignado_por) "
        "SELECT u.id, r.id, NULL FROM usuarios u JOIN roles r ON r.codigo = 'docente' "
        "WHERE u.email IN ("
        + ", ".join(f"'{teacher['email']}'" for teacher in payload["teachers"]) +
        ") ON CONFLICT (usuario_id, rol_id) DO NOTHING;"
    )
    lines.append("")

    lines.append("-- Asignaturas")
    for subject in payload["subjects"]:
        lines.append(
            "INSERT INTO asignaturas (programa_id, codigo, nombre, creditos, activa) "
            "SELECT id, '{codigo}', '{nombre}', 3, true FROM programas_academicos WHERE codigo = 'ING-SIS' "
            "ON CONFLICT (programa_id, codigo) DO NOTHING;".format(
                codigo=subject["codigo"],
                nombre=subject["nombre"].replace("'", "''"),
            )
        )
    lines.append("")

    lines.append("-- Docentes")
    for teacher in payload["teachers"]:
        lines.append(
            "INSERT INTO docentes (usuario_id, codigo_docente, activo) "
            "SELECT id, '{codigo_docente}', true FROM usuarios WHERE email = '{email}' "
            "ON CONFLICT (usuario_id) DO NOTHING;".format(
                codigo_docente=re.sub(r"[^a-z0-9]+", "", teacher["email"].split("@", 1)[0]).lower(),
                email=teacher["email"],
            )
        )
    lines.append("")

    lines.append("-- Grupos de asignatura")
    for assignment in payload["assignments"]:
        subject = next(subject for subject in payload["subjects"] if canonical_key(subject["nombre"]) == canonical_key(assignment["subject_name"]))
        lines.append(
            "INSERT INTO grupos_asignatura (asignatura_id, docente_id, codigo_grupo, jornada, activo) "
            "SELECT a.id, d.id, '{grupo}', '{jornada}', true FROM asignaturas a "
            "JOIN programas_academicos p ON p.id = a.programa_id "
            "JOIN docentes d ON d.usuario_id = (SELECT u.id FROM usuarios u WHERE u.email = '{email}') "
            "WHERE p.codigo = 'ING-SIS' AND a.codigo = '{codigo}' "
            "ON CONFLICT (asignatura_id, codigo_grupo) DO NOTHING;".format(
                grupo=assignment["group_code"],
                jornada=assignment["jornada"],
                email=assignment["email"],
                codigo=subject["codigo"],
            )
        )
    lines.append("")
    lines.append("COMMIT;")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    excel_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not excel_path.exists():
        raise FileNotFoundError(f"No se encontró el Excel: {excel_path}")

    sql_output_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_SQL_OUTPUT

    rows = read_rows(excel_path)
    payload = build_seed_payload(rows)
    write_sql_output(payload, sql_output_path)

    print(f"Filas leídas del Excel: {len(rows)}")
    print(f"Docentes únicos: {len(payload['teachers'])}")
    print(f"Asignaturas únicas: {len(payload['subjects'])}")
    print(f"Vínculos de grupos: {len(payload['assignments'])}")
    print(f"SQL generado en: {sql_output_path}")


if __name__ == "__main__":
    main()
