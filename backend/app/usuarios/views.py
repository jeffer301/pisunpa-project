import re
from uuid import UUID

from django.db import models, transaction
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.permissions import (
    AllowAny,
    IsAuthenticated,
)
from rest_framework import generics
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView

from app.egresados.models import PerfilEgresado, Programa
from .models import InvitacionDocente, Notificacion, Rol, Usuario
from .notification_service import NotificacionService
from .permissions import EsAdminEscritura, EsAdminLectura
from .serializers import (
    CambioRolSerializer,
    CrearAdminSerializer,
    CustomTokenObtainSerializer,
    EstudiantePendienteSerializer,
    InvitacionDetailSerializer,
    InvitacionDocenteSerializer,
    NotificacionSerializer,
    ProfesorSerializer,
    RegistroConRolSerializer,
    RegistroDocenteConTokenSerializer,
    RegistroDocenteSerializer,
    RegistroSerializer,
    UsuariosDisponiblesSerializer,
    UsuarioGestionSerializer,
    UsuarioSerializer,
)
from .utils import enviar_invitacion_docente


class RegistroConRolView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegistroConRolSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        with transaction.atomic():
            usuario = Usuario(
                username=data["email"],
                email=data["email"],
                first_name=data["first_name"],
                last_name=data["last_name"],
                documento=data["documento"],
                documento_identidad=data.get("documento_identidad", ""),
                telefono=data.get("telefono", ""),
            )
            usuario.set_password(data["password"])
            usuario.estado = "pendiente_aprobacion"
            usuario.save()

            if data["tipo_usuario"] == "estudiante":
                usuario.rol = Rol.objects.get(nombre='estudiante')
                usuario.save(update_fields=['rol'])

            if data["tipo_usuario"] == "egresado":
                programa = Programa.objects.get(
                    id=data["programa_id"]
                )
                PerfilEgresado.objects.create(
                    usuario=usuario,
                    tipo_documento="CC",
                    numero_documento=data["documento"],
                    telefono_celular=data.get("telefono", ""),
                    direccion_residencia=data.get(
                        "direccion_residencia", ""
                    ),
                    biografia=data.get("biografia", ""),
                    programa=programa,
                    validado=False,
                )

        return Response(
            {
                "mensaje": (
                    "Registro como egresado pendiente de "
                    "validaciÃ³n."
                    if data["tipo_usuario"] == "egresado"
                    else "Tu cuenta estÃ¡ pendiente de aprobaciÃ³n "
                    "por el director/administrador."
                )
            },
            status=status.HTTP_201_CREATED,
        )


class RegistroView(generics.CreateAPIView):
    queryset = Usuario.objects.all()
    serializer_class = RegistroSerializer
    permission_classes = [AllowAny]


class PerfilView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UsuarioSerializer(request.user)
        return Response(serializer.data)

    def put(self, request):
        serializer = UsuarioSerializer(
            request.user,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainSerializer


class UsuariosDisponiblesView(generics.ListAPIView):
    serializer_class = UsuariosDisponiblesSerializer
    permission_classes = [IsAuthenticated, EsAdminLectura]

    def get_queryset(self):
        usuario_ids_con_perfil = PerfilEgresado.objects.values_list(
            'usuario_id', flat=True
        )
        return Usuario.objects.exclude(id__in=usuario_ids_con_perfil)


class RegistroDocenteView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegistroDocenteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        with transaction.atomic():
            usuario = Usuario(
                username=data["email"],
                email=data["email"],
                first_name=data["first_name"],
                last_name=data["last_name"],
                documento=data["documento_identidad"],
                documento_identidad=data["documento_identidad"],
            )
            usuario.set_password(data["password"])
            usuario.rol = Rol.objects.get(nombre='profesor')
            usuario.estado = 'aprobado'
            usuario.save()

        return Response(
            {"mensaje": "Registro exitoso. Ya puedes iniciar sesiÃ³n."},
            status=status.HTTP_201_CREATED,
        )


class EstudiantesPendientesView(generics.ListAPIView):
    serializer_class = EstudiantePendienteSerializer
    permission_classes = [IsAuthenticated, EsAdminLectura]

    def get_queryset(self):
        return Usuario.objects.filter(
            estado='pendiente_aprobacion'
        ).select_related('rol').order_by('-creado')


class AprobarEstudianteView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def patch(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk)
        except Usuario.DoesNotExist:
            return Response(
                {"error": "Usuario no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )
        usuario.estado = 'aprobado'
        usuario.save()
        return Response({"mensaje": "Estudiante aprobado correctamente"})


class RechazarEstudianteView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def patch(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk)
        except Usuario.DoesNotExist:
            return Response(
                {"error": "Usuario no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )
        usuario.estado = 'rechazado'
        usuario.save()
        return Response({"mensaje": "Estudiante rechazado"})


class PromoverEgresadoView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def patch(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk)
        except Usuario.DoesNotExist:
            return Response(
                {"error": "Usuario no encontrado"},
                status=status.HTTP_404_NOT_FOUND
            )

        from app.usuarios.models import Rol
        from app.egresados.models import PerfilEgresado

        rol_egresado = Rol.objects.get(nombre='egresado')

        with transaction.atomic():
            usuario.rol = rol_egresado
            usuario.estado = 'aprobado'
            usuario.save(update_fields=['rol', 'estado'])

            if not hasattr(usuario, 'perfil_egresado'):
                programa_id = request.data.get('programa_id')
                programa = None
                if programa_id:
                    from app.egresados.models import Programa
                    try:
                        programa = Programa.objects.get(id=programa_id)
                    except Programa.DoesNotExist:
                        pass

                PerfilEgresado.objects.create(
                    usuario=usuario,
                    tipo_documento='CC',
                    numero_documento=usuario.documento or f'promo_{usuario.id}',
                    programa=programa,
                    validado=True,
                )

        return Response({"mensaje": "Usuario promovido a egresado correctamente"})


class NotificacionListView(generics.ListAPIView):
    serializer_class = NotificacionSerializer

    def get_queryset(self):
        return NotificacionService.obtener_notificaciones(self.request.user)


class NotificacionContarNoLeidasView(APIView):
    def get(self, request):
        count = NotificacionService.contar_no_leidas(request.user)
        return Response({'count': count})


class NotificacionMarcarLeidaView(APIView):
    def patch(self, request, pk):
        updated = NotificacionService.marcar_como_leida(pk, request.user)
        if not updated:
            return Response({'detail': 'NotificaciÃ³n no encontrada'}, status=404)
        return Response({'detail': 'Marcada como leÃ­da'})


class NotificacionLeerTodasView(APIView):
    def post(self, request):
        Notificacion.objects.filter(usuario=request.user, leido=False).update(leido=True)
        return Response({'detail': 'Todas marcadas como leÃ­das'})


class ProfesorListView(generics.ListAPIView):
    serializer_class = ProfesorSerializer
    permission_classes = [IsAuthenticated, EsAdminLectura]

    def get_queryset(self):
        return Usuario.objects.filter(
            rol__nombre='profesor'
        ).select_related('rol').prefetch_related('invitaciones').order_by('-creado')


class ProfesorImportView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'Debe subir un archivo Excel.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(archivo, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=3, values_only=True))
        except Exception:
            return Response({'error': 'Formato de archivo no vÃ¡lido.'}, status=status.HTTP_400_BAD_REQUEST)

        rol_profesor = Rol.objects.get(nombre='profesor')
        creados = 0
        duplicados = 0
        errores = []
        profesores_unicos = {}

        for row in rows:
            nombre = str(row[0]).strip() if row[0] else None
            email = str(row[1]).strip() if row[1] else None

            if not nombre or not email:
                continue

            email = email.lower()
            email_ok = re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email)

            if not email_ok:
                errores.append(f'Email invÃ¡lido para {nombre}: {row[1]}')
                continue

            if email in profesores_unicos:
                continue

            profesores_unicos[email] = nombre

        for email, nombre in profesores_unicos.items():
            if Usuario.objects.filter(email=email).exists():
                duplicados += 1
                continue

            partes = nombre.split(' ', 1)
            first_name = partes[0]
            last_name = partes[1] if len(partes) > 1 else ''

            with transaction.atomic():
                usuario = Usuario(
                    username=email,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    documento=email,
                    rol=rol_profesor,
                    estado='pendiente_aprobacion',
                )
                usuario.set_password(Usuario.objects.make_random_password())
                usuario.save()
                creados += 1

        return Response({
            'creados': creados,
            'duplicados': duplicados,
            'errores': errores,
            'total_procesados': len(profesores_unicos),
        })


class ProfesorInvitarView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def post(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk, rol__nombre='profesor')
        except Usuario.DoesNotExist:
            return Response({'error': 'Profesor no encontrado.'}, status=404)

        invitacion_valida = InvitacionDocente.objects.filter(
            usuario=usuario, usado=False
        ).first()

        if invitacion_valida and not invitacion_valida.valido:
            invitacion_valida = None

        if invitacion_valida:
            token = str(invitacion_valida.token)
            email = invitacion_valida.email
            creada_ahora = False
        else:
            invitacion = InvitacionDocente.objects.create(
                usuario=usuario,
                email=usuario.email,
            )
            token = str(invitacion.token)
            email = invitacion.email
            creada_ahora = True

        nombre = f'{usuario.first_name} {usuario.last_name}'.strip() or email
        enviar_invitacion_docente(email=email, nombre=nombre, token=token)

        if creada_ahora:
            serializer = InvitacionDocenteSerializer(invitacion)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            serializer = InvitacionDocenteSerializer(invitacion_valida)
            return Response(serializer.data)


class InvitacionDetailView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, token):
        try:
            token_uuid = UUID(str(token))
            invitacion = InvitacionDocente.objects.select_related('usuario').get(token=token_uuid)
        except (ValueError, InvitacionDocente.DoesNotExist):
            return Response({'error': 'El enlace de invitaciÃ³n no es vÃ¡lido.'}, status=404)

        if not invitacion.valido:
            return Response({'error': 'La invitaciÃ³n ha expirado o ya fue utilizada.'}, status=410)

        serializer = InvitacionDetailSerializer(data={
            'email': invitacion.email,
            'first_name': invitacion.usuario.first_name,
            'last_name': invitacion.usuario.last_name,
            'token': invitacion.token,
            'valido': invitacion.valido,
        })
        serializer.is_valid()
        return Response(serializer.data)


class RegistroDocenteConTokenView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegistroDocenteConTokenSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        with transaction.atomic():
            invitacion = InvitacionDocente.objects.select_related('usuario').get(token=data['token'])
            usuario = invitacion.usuario

            usuario.first_name = data['first_name']
            usuario.last_name = data['last_name']
            usuario.documento = data['documento_identidad']
            usuario.documento_identidad = data['documento_identidad']
            usuario.estado = 'aprobado'
            usuario.set_password(data['password'])
            usuario.save()

            invitacion.usado = True
            invitacion.save()

        return Response(
            {'mensaje': 'Registro exitoso. Ya puedes iniciar sesión.'},
            status=status.HTTP_201_CREATED,
        )


class GestUsuariosListView(generics.ListAPIView):
    serializer_class = UsuarioGestionSerializer
    permission_classes = [IsAuthenticated, EsAdminLectura]

    def get_queryset(self):
        qs = Usuario.objects.select_related('rol').order_by('-creado')
        query = self.request.query_params.get('q', '').strip()
        estado = self.request.query_params.get('estado', '').strip()
        if query:
            qs = qs.filter(
                models.Q(email__icontains=query)
                | models.Q(first_name__icontains=query)
                | models.Q(last_name__icontains=query)
                | models.Q(documento__icontains=query)
            )
        if estado:
            qs = qs.filter(estado=estado)
        return qs


class CambiarRolView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def patch(self, request, pk):
        try:
            usuario = Usuario.objects.get(pk=pk)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado.'}, status=404)
        serializer = CambioRolSerializer(
            data=request.data,
            context={'usuario': usuario, 'solicitante': request.user},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UsuarioGestionSerializer(usuario).data)


class CrearAdminView(APIView):
    permission_classes = [IsAuthenticated, EsAdminEscritura]

    def post(self, request):
        serializer = CrearAdminSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        usuario = serializer.save()
        return Response(
            UsuarioGestionSerializer(usuario).data,
            status=status.HTTP_201_CREATED,
        )
