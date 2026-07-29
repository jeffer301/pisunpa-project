import re

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from app.usuarios.models import Rol, Usuario


class Command(BaseCommand):
    help = 'Importa profesores desde un archivo Excel institucional'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo Excel')
        parser.add_argument(
            '--fila-inicio', type=int, default=3,
            help='Fila donde comienzan los datos (default: 3)',
        )

    def handle(self, *args, **options):
        ruta = options['archivo']
        fila_inicio = options['fila_inicio']

        try:
            wb = load_workbook(ruta, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=fila_inicio, values_only=True))
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {ruta}')
        except Exception as e:
            raise CommandError(f'Error al leer el archivo: {e}')

        if not rows:
            raise CommandError('El archivo Excel está vacío.')

        rol_profesor = Rol.objects.get(nombre='profesor')
        creados = 0
        duplicados = 0
        errores = []
        profesores_unicos = {}

        for row in rows:
            nombre = str(row[0]).strip() if row[0] else None
            email = str(row[1]).strip().lower() if row[1] else None

            if not nombre or not email:
                continue

            email_ok = re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email)
            if not email_ok:
                errores.append(f'Email inválido para "{nombre}": {row[1]}')
                continue

            if email in profesores_unicos:
                continue

            profesores_unicos[email] = nombre

        for email, nombre in profesores_unicos.items():
            if Usuario.objects.filter(email=email).exists():
                duplicados += 1
                self.stdout.write(f'  = {email} ya existe (duplicado)')
                continue

            partes = nombre.split(' ', 1)
            first_name = partes[0]
            last_name = partes[1] if len(partes) > 1 else ''

            usuario = Usuario(
                username=email,
                email=email,
                first_name=first_name,
                last_name=last_name,
                documento=email,
                rol=rol_profesor,
                estado='pendiente_aprobacion',
            )
            usuario.set_password(Usuario.objects.make_random_password())
            usuario.save()
            creados += 1
            self.stdout.write(f'  + {email} ({nombre})')

        self.stdout.write()
        self.stdout.write(self.style.SUCCESS(
            f'Profesores: {creados} creados, {duplicados} duplicados, '
            f'{len(errores)} errores (de {len(profesores_unicos)} únicos)'
        ))
        if errores:
            self.stdout.write(self.style.WARNING('Errores:'))
            for e in errores:
                self.stdout.write(f'  - {e}')
