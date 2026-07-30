import re
import unicodedata
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from django.contrib.auth import get_user_model
from app.egresados.models import Asignatura, Grupo, ProfesorAsignatura, Programa

User = get_user_model()


def normalize(text: str) -> str:
    text = unicodedata.normalize('NFD', text)
    text = text.encode('ascii', 'ignore').decode('utf-8')
    return text.strip().upper()


class Command(BaseCommand):
    help = 'Importa grupos desde el Excel institucional (mismo archivo de profesores)'

    def add_arguments(self, parser):
        parser.add_argument('archivo', type=str, help='Ruta al archivo Excel')
        parser.add_argument('--fila-inicio', type=int, default=3)

    def handle(self, *args, **options):
        ruta = options['archivo']
        fila_inicio = options['fila_inicio']

        try:
            wb = load_workbook(ruta, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=fila_inicio, values_only=True))
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {ruta}')
        except Exception as e:
            raise CommandError(f'Error al leer el archivo: {e}')

        if not rows:
            raise CommandError('El archivo Excel está vacío.')

        programa = Programa.objects.filter(nombre__icontains='Ingenier').first()
        if not programa:
            raise CommandError('No se encontró el programa "Ingeniería de Sistemas". Ejecute primero seed_asignaturas.')

        creados = 0
        existentes = 0
        sin_profesor = 0
        profesor_actual = None
        errores = []

        for row in rows:
            nombre_profesor = str(row[0]).strip() if row[0] else None
            email_profesor = str(row[1]).strip().lower() if row[1] else None
            nombre_asignatura = str(row[3]).strip() if row[3] else None
            codigo_grupo = str(row[4]).strip() if row[4] else None
            franja = str(row[5]).strip() if row[5] else None

            if nombre_profesor and email_profesor:
                email_ok = re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_profesor)
                if email_ok:
                    profesor_actual = User.objects.filter(email=email_profesor).first()
                    if not profesor_actual:
                        errores.append(f'Profesor no encontrado en DB: {nombre_profesor} <{email_profesor}>')

            if not nombre_asignatura or not codigo_grupo:
                continue

            asignatura = self._get_or_create_asignatura(nombre_asignatura, programa)

            grupo, created = Grupo.objects.get_or_create(
                codigo=codigo_grupo.upper(),
                asignatura=asignatura,
                defaults={
                    'profesor': profesor_actual,
                    'franja_horaria': franja or '',
                }
            )
            if created:
                creados += 1
            else:
                existentes += 1

            if profesor_actual:
                ProfesorAsignatura.objects.get_or_create(
                    profesor=profesor_actual,
                    asignatura=asignatura,
                )

        self.stdout.write()
        self.stdout.write(self.style.SUCCESS(
            f'Grupos: {creados} creados, {existentes} ya existentes'
        ))
        if sin_profesor:
            self.stdout.write(self.style.WARNING(
                f'{sin_profesor} grupos sin profesor asignado'
            ))
        if errores:
            self.stdout.write(self.style.WARNING('Advertencias:'))
            for e in errores:
                self.stdout.write(f'  - {e}')

    def _get_or_create_asignatura(self, nombre: str, programa) -> Asignatura:
        normalizado = normalize(nombre)
        asignatura = Asignatura.objects.filter(nombre__iexact=normalizado).first()
        if not asignatura:
            asignatura = Asignatura.objects.create(
                nombre=normalizado,
                programa=programa,
            )
            self.stdout.write(f'  + Asignatura creada: {normalizado}')
        return asignatura
